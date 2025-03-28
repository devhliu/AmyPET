"""
Utility functions for centiloid workflow.

This module provides utility functions for the centiloid workflow.
"""
import os
import pickle
import numpy as np
import openpyxl as xl
from pathlib import Path

def get_cl_anchors(path):
    """
    Load centiloid anchor values from a file.
    
    Args:
        path: Path to the anchor file or directory
        
    Returns:
        Dictionary with anchor values
    """
    from amypet.utils import get_cl_anchors as _get_cl_anchors
    return _get_cl_anchors(path)

def get_ur2pib(tracer, path=None):
    """
    Get uptake ratio to PiB conversion factors.
    
    Args:
        tracer: Tracer name ('fbb', 'fbp', 'flute')
        path: Path to the conversion file or directory
        
    Returns:
        Dictionary with conversion factors
    """
    from amypet.utils import get_ur2pib as _get_ur2pib
    return _get_ur2pib(tracer, path=path)

def get_clref(excel_file):
    """
    Get reference centiloid values from an Excel file.
    
    Args:
        excel_file: Path to Excel file with reference data
        
    Returns:
        Dictionary with reference values
    """
    # Load workbook
    wb = xl.load_workbook(excel_file)
    ws = wb.active
    
    pib_tbl = {'yc': {}, 'ad': {}}
    
    # Cell offsets
    ioff_ad = [5, 50]
    ioff_yc = [51, 85]
    
    # Extract data from columns
    sbj = [str(i.value) for i in ws['A']]
    cg = [str(i.value) for i in ws['B']]
    wc = [str(i.value) for i in ws['C']]
    wcb = [str(i.value) for i in ws['D']]
    pns = [str(i.value) for i in ws['E']]
    
    ccg = [str(i.value) for i in ws['F']]
    cwc = [str(i.value) for i in ws['G']]
    cwcb = [str(i.value) for i in ws['H']]
    cpns = [str(i.value) for i in ws['I']]
    
    # Young control indices and UR/CL values
    pib_tbl['yc']['id'] = np.array([int(i[3:]) for i in sbj[ioff_yc[0]:ioff_yc[1]]])
    
    pib_tbl['yc']['ur'] = {}
    pib_tbl['yc']['ur']['cg'] = np.array([float(i) for i in cg[ioff_yc[0]:ioff_yc[1]]])
    pib_tbl['yc']['ur']['wc'] = np.array([float(i) for i in wc[ioff_yc[0]:ioff_yc[1]]])
    pib_tbl['yc']['ur']['wcb'] = np.array([float(i) for i in wcb[ioff_yc[0]:ioff_yc[1]]])
    pib_tbl['yc']['ur']['pns'] = np.array([float(i) for i in pns[ioff_yc[0]:ioff_yc[1]]])
    
    pib_tbl['yc']['cl'] = {}
    pib_tbl['yc']['cl']['cg'] = np.array([float(i) for i in ccg[ioff_yc[0]:ioff_yc[1]]])
    pib_tbl['yc']['cl']['wc'] = np.array([float(i) for i in cwc[ioff_yc[0]:ioff_yc[1]]])
    pib_tbl['yc']['cl']['wcb'] = np.array([float(i) for i in cwcb[ioff_yc[0]:ioff_yc[1]]])
    pib_tbl['yc']['cl']['pns'] = np.array([float(i) for i in cpns[ioff_yc[0]:ioff_yc[1]]])
    
    # AD indices and UR/CL values
    pib_tbl['ad']['id'] = np.array([int(i[3:]) for i in sbj[ioff_ad[0]:ioff_ad[1]]])
    
    pib_tbl['ad']['ur'] = {}
    pib_tbl['ad']['ur']['cg'] = np.array([float(i) for i in cg[ioff_ad[0]:ioff_ad[1]]])
    pib_tbl['ad']['ur']['wc'] = np.array([float(i) for i in wc[ioff_ad[0]:ioff_ad[1]]])
    pib_tbl['ad']['ur']['wcb'] = np.array([float(i) for i in wcb[ioff_ad[0]:ioff_ad[1]]])
    pib_tbl['ad']['ur']['pns'] = np.array([float(i) for i in pns[ioff_ad[0]:ioff_ad[1]]])
    
    pib_tbl['ad']['cl'] = {}
    pib_tbl['ad']['cl']['cg'] = np.array([float(i) for i in ccg[ioff_ad[0]:ioff_ad[1]]])
    pib_tbl['ad']['cl']['wc'] = np.array([float(i) for i in cwc[ioff_ad[0]:ioff_ad[1]]])
    pib_tbl['ad']['cl']['wcb'] = np.array([float(i) for i in cwcb[ioff_ad[0]:ioff_ad[1]]])
    pib_tbl['ad']['cl']['pns'] = np.array([float(i) for i in cpns[ioff_ad[0]:ioff_ad[1]]])
    
    return pib_tbl

def check_urs(out_yc, out_ad, refs):
    """
    Check uptake ratios against reference values.
    
    Args:
        out_yc: Young controls output
        out_ad: AD patients output
        refs: Reference values
        
    Returns:
        Dictionary with comparison results
    """
    diff = {'yc': {}, 'ad': {}}
    
    # Process each reference region
    for rvoi in ['cg', 'wc', 'wcb', 'pns']:
        diff['yc'][rvoi] = {}
        diff['ad'][rvoi] = {}
        
        # Process young controls
        for i, sbj in enumerate(out_yc['sbj']):
            sbj_id = int(sbj)
            idx = np.where(refs['yc']['id'] == sbj_id)[0]
            if len(idx) > 0:
                idx = idx[0]
                diff['yc'][rvoi][f'YC-{sbj_id}'] = {
                    'ur': out_yc[rvoi][i],
                    'ur_ref': refs['yc']['ur'][rvoi][idx],
                    'cl_ref': refs['yc']['cl'][rvoi][idx]
                }
        
        # Process AD patients
        for i, sbj in enumerate(out_ad['sbj']):
            sbj_id = int(sbj)
            idx = np.where(refs['ad']['id'] == sbj_id)[0]
            if len(idx) > 0:
                idx = idx[0]
                diff['ad'][rvoi][f'AD-{sbj_id}'] = {
                    'ur': out_ad[rvoi][i],
                    'ur_ref': refs['ad']['ur'][rvoi][idx],
                    'cl_ref': refs['ad']['cl'][rvoi][idx]
                }
    
    return diff

def check_cls(out_yc, out_ad, diff, refs):
    """
    Calculate centiloid values and compare with reference values.
    
    Args:
        out_yc: Young controls output
        out_ad: AD patients output
        diff: Comparison results from check_urs
        refs: Reference values
        
    Returns:
        Updated comparison results with centiloid values
    """
    # Process each reference region
    for rvoi in ['cg', 'wc', 'wcb', 'pns']:
        # Get YC mean uptake ratio
        yc_ur_mean = np.mean([diff['yc'][rvoi][k]['ur'] for k in diff['yc'][rvoi]])
        
        # Get AD mean uptake ratio
        ad_ur_mean = np.mean([diff['ad'][rvoi][k]['ur'] for k in diff['ad'][rvoi]])
        
        # Calculate centiloid values for young controls
        for k in diff['yc'][rvoi]:
            ur = diff['yc'][rvoi][k]['ur']
            diff['yc'][rvoi][k]['cl'] = (ur - yc_ur_mean) / (ad_ur_mean - yc_ur_mean) * 100
        
        # Calculate centiloid values for AD patients
        for k in diff['ad'][rvoi]:
            ur = diff['ad'][rvoi][k]['ur']
            diff['ad'][rvoi][k]['cl'] = (ur - yc_ur_mean) / (ad_ur_mean - yc_ur_mean) * 100
    
    return diff

def save_cl_anchors(diff, outpath=None):
    """
    Save centiloid anchor values.
    
    Args:
        diff: Comparison results from check_cls
        outpath: Output directory
        
    Returns:
        Dictionary with anchor values
    """
    # Create anchor dictionary
    anchors = {}
    
    # Process each reference region
    for rvoi in ['cg', 'wc', 'wcb', 'pns']:
        # Get YC mean uptake ratio
        yc_ur_mean = np.mean([diff['yc'][rvoi][k]['ur'] for k in diff['yc'][rvoi]])
        
        # Get AD mean uptake ratio
        ad_ur_mean = np.mean([diff['ad'][rvoi][k]['ur'] for k in diff['ad'][rvoi]])
        
        # Store anchor values
        anchors[rvoi] = {
            'yc_ur': yc_ur_mean,
            'ad_ur': ad_ur_mean
        }
    
    # Save anchors if outpath is provided
    if outpath:
        outpath = Path(outpath)
        os.makedirs(outpath, exist_ok=True)
        
        with open(outpath / 'cl_anchors.pkl', 'wb') as f:
            pickle.dump(anchors, f)
    
    return anchors